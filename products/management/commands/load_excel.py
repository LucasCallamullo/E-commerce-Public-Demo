from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils.text import slugify
from django.utils.crypto import get_random_string

from openpyxl import load_workbook
from django.contrib.postgres.search import SearchVector

from core.utils.utils_parsers import normalize_or_None

from products.models.product import Product
from products.models.category import Category
from products.models.subcategory import Subcategory
from products.models.brand import Brand
from products.models.product_image import ProductImage


def update_defaults_slugs():
    # Categorías
    categories_to_update = []
    for c in Category.objects.filter(is_default=False):
        if not c.slug:
            c.slug = slugify(c.name)
            categories_to_update.append(c)
    if categories_to_update:
        Category.objects.bulk_update(categories_to_update, ['slug'])

    # Subcategorías
    subcategories_to_update = []
    for sc in Subcategory.objects.filter(is_default=False):
        if not sc.slug:
            sc.slug = slugify(sc.name)
            subcategories_to_update.append(sc)
    if subcategories_to_update:
        Subcategory.objects.bulk_update(subcategories_to_update, ['slug'])

    # Marcas
    brands_to_update = []
    for b in Brand.objects.filter(is_default=False):
        if not b.slug:
            b.slug = slugify(b.name)
            brands_to_update.append(b)
    if brands_to_update:
        Brand.objects.bulk_update(brands_to_update, ['slug'])

    print("Slugs actualizados correctamente.")
    

# command python manage.py load_data_project
def unique_slug(base_slug, model):
    slug = base_slug
    while model.objects.filter(slug=slug).exists():
        slug = f"{base_slug}-{get_random_string(4)}"
    return slug


def clean_value(value, zero=False):
    """
        Cleans null or empty values and converts them to None or Zero as appropriate 
        for the database and stores them correctly.

    Args:
        value (any): The value obtained from the data source (e.g., an Excel file).
        zero (bool): Enables returning 0 instead of None for numeric fields.

    Returns:
        Returns a valid value or None or Zero as appropriate
    """
    result = None if not value or value == '' else value
    if zero:
        return 0 if result is None else result
    
    return result


class Command(BaseCommand):
    help = "To generically load all the necessary data for the models we created "
    help += "in this case, it includes loading Product models from Excel to the database. "
    help += "On the other hand, dictionaries were simply used to create model examples like Store, User, Orders"
        
    def handle(self, *args, **kwargs):
        self.load_products_from_excel()
        # update_defaults_slugs()
        # self.update_main_imagess()
        

    def load_products_from_excel(self, file_path='products/data/products_data.xlsx'):
        """
        Importa productos desde un archivo Excel a la base de datos.

        - Crea categorías, subcategorías y marcas si no existen.
        - Actualiza productos existentes según el nombre.
        - Añade imágenes asociadas evitando duplicados.
        """

        try:
            wb = load_workbook(file_path)
        except FileNotFoundError:
            print(f"Archivo no encontrado: {file_path}")
            return

            
        qs = Product.objects.all()
        if qs.exists():
            print("Ya se cargo el excel")
            return
        
        """ 
        # Primero borramos las imágenes asociadas
        ProductImage.objects.all().delete()

        # Luego borramos los productos
        Product.objects.all().delete()
        """
    
        ws = wb.active
        

        # Columnas esperadas
        columns = [
            'id', 'name', 'price', 'available', 'stock', 'category', 'subcategory', 'brand',
            'discount', 'description', 'image_url', 'image_url2'
        ]

        # Cache para no repetir queries
        categories_cache = {}
        subcategories_cache = {}
        brands_cache = {}

        # Pre-cargar los defaults
        default_category = Category.objects.filter(is_default=True, id=1).first()
        # default_subcategory = Subcategory.objects.filter(is_default=True, id=1).first()
        default_subcategory = None
        default_brand = Brand.objects.filter(is_default=True, id=1).first()

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(columns, row))

            # Nombre del producto
            name = clean_value(row_dict.get("name"))
            if not name:
                fila_num = int(row_dict.get("id", 1)) + 1
                print(f'Fila {fila_num}: producto sin nombre, se ignora.')
                continue

            # Categoría
            category_value = clean_value(row_dict.get("category"))
            if not category_value:
                category = default_category
            else:
                if category_value in categories_cache:
                    category = categories_cache[category_value]
                else:
                    slug_cat = slugify(category_value)
                    category, _ = Category.objects.get_or_create(name=category_value, slug=slug_cat)
                    categories_cache[category_value] = category

            # Subcategoría
            subcategory_value = clean_value(row_dict.get("subcategory"))
            subcat_key = (subcategory_value, category.id)
            if not subcategory_value:
                subcategory = default_subcategory
            else:
                if subcat_key in subcategories_cache:
                    subcategory = subcategories_cache[subcat_key]
                else:
                    slug_subcat = slugify(subcategory_value)
                    subcategory, _ = Subcategory.objects.get_or_create(
                        name=subcategory_value,
                        slug=slug_subcat,
                        category=category
                    )
                    subcategories_cache[subcat_key] = subcategory

            # Marca
            brand_value = clean_value(row_dict.get("brand"))
            if not brand_value:
                brand = default_brand
            else:
                if brand_value in brands_cache:
                    brand = brands_cache[brand_value]
                else:
                    slug_brand = slugify(brand_value)
                    brand, _ = Brand.objects.get_or_create(name=brand_value, slug=slug_brand)
                    brands_cache[brand_value] = brand

            # Otros campos
            price = clean_value(row_dict.get("price"), zero=True)
            stock = clean_value(row_dict.get("stock"), zero=True)
            discount = clean_value(row_dict.get("discount"), zero=True)
            available = str(row_dict.get("available", "")).strip().lower() in ['si', 'sí', 'yes', '1']
            description = clean_value(row_dict.get("description"))

            normalized_name = normalize_or_None(name)
            slug = unique_slug(slugify(name), Product)

            # Guardamos producto (update_or_create para actualizar si ya existe)
            with transaction.atomic():
                product_obj, created = Product.objects.update_or_create(
                    name=name,
                    defaults={
                        'slug': slug,
                        'normalized_name': normalized_name,
                        'price_ars': price,
                        'stock': stock,
                        'discount_ars': discount,
                        'available': available,
                        'category': category,
                        'subcategory': subcategory,
                        'brand': brand,
                        'description': description,
                    }
                )

                # Manejo de imágenes (evita duplicados)
                image_url = clean_value(row_dict.get("image_url"))
                image_url2 = clean_value(row_dict.get("image_url2"))

                cont = 0
                for url in filter(None, [image_url, image_url2]):
                    if not url:
                        continue
                    
                    if not ProductImage.objects.filter(product=product_obj, image_url=url).exists():
                        ProductImage.objects.create(product=product_obj, image_url=url)
                        cont += 1

            if created:
                print(f'Product {product_obj.name} created successfully with {cont} associated images.')
            else:
                print(f'Product {product_obj.name} updated successfully with {cont} associated images.')

        # por como es el flujo para los search vectors necesitamos llamarlo despues de haber creado los productos
        # entonces hacemos el update en bulk posterior
        self.update_normalized_name()
    
        # despues actualizar las main_image para evitar consulta extra
        self.update_main_imagess()
    
    def update_normalized_name(self):
        products = []
        for product in Product.objects.all():
            if product.name:
                product.normalized_name = normalize_or_None(product.name)
                products.append(product)

        Product.objects.bulk_update(products, ['normalized_name'])
        Product.objects.update(search_vector=SearchVector('normalized_name', weight='A'))
        
        
    def update_main_imagess(self):
        products = Product.objects.all()

        for product in products:
            images = product.images.all()
            
            if not images:
                print(f'Product "{product.name}" has no images.')
                continue

            # Set the first image as main
            first_image = images[0]
            first_image.main_image = True
            first_image.save()

            # Optionally, set all others to False (cleanup)
            for img in images[1:]:
                if img.main_image:
                    img.main_image = False
                    img.save()

            # Update product's main_image field
            product.main_image = first_image.image_url
            product.save()
            print(f'✔ Main image updated for product "{product.name}"')
            
    